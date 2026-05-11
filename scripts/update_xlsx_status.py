"""
Update links_v2.xlsx Sheet1 with a status column (F = col 6):
  T = scraper implemented and working
  F = probed, known to fail (JS-rendered, SSL, 404, blocked, no tables, etc.)
  blank = not yet probed / unclassified

Run from repo root:
  python scripts/update_xlsx_status.py
"""
from __future__ import annotations
import importlib
import sys
import re
from pathlib import Path
from urllib.parse import urlparse

# ensure repo root is on the path so `scrapers` package is importable
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# ── collect working source_urls from all registered scrapers ──────────────
SCRAPER_MODULES = [
    "scrapers.chungcheongnam.asan",
    "scrapers.chungcheongnam.chungnam_batch",
    "scrapers.chungcheongbuk.chungbuk_batch",
    "scrapers.gyeonggi.gyeonggi_batch",
    "scrapers.jeonbuk.jeonbuk_batch",
    "scrapers.jeonbuk.jeonbuk_list_batch",
    "scrapers.jeonnam.jeonnam_batch",
    "scrapers.jeonnam.jeonnam_list_batch",
    "scrapers.daejeon.daejeon_si",
    "scrapers.seoul.seoul_si",
    "scrapers.seoul.suwon_kukto",
    "scrapers.seoul.sisul",
    "scrapers.seoul.ish",
    "scrapers.seoul.doro_seoul",
    "scrapers.busan.busan_si",
    "scrapers.busan.jinju_kukto",
    "scrapers.busan.daegu_kukto",
    "scrapers.busan.pohang_kukto",
    "scrapers.busan.yeongju_kukto",
    "scrapers.busan.busan_gunsul",
    "scrapers.busan.bisco",
    "scrapers.busan.busan_batch",
    "scrapers.daegu.daegu_dosi",
    "scrapers.incheon.incheon_si",
    "scrapers.incheon.jonggeon",
    "scrapers.gwangju.gmcc",
    "scrapers.daejeon.daejeon_kukto",
    "scrapers.daejeon.nonsan_kukto",
    "scrapers.daejeon.chungju_kukto",
    "scrapers.daejeon.boeun_kukto",
    "scrapers.daejeon.yesan_kukto",
    "scrapers.daejeon.daejeon_gunseol",
    "scrapers.daejeon.dcco",
    "scrapers.ulsan.umca",
    "scrapers.ulsan.ulsan_si",
    "scrapers.sejong.sejong_si_notice",
    "scrapers.sejong.sejong_si_gosi",
    "scrapers.sejong.sejong_batch",
    "scrapers.gyeongsangbuk.gyeongbuk_batch",
    "scrapers.gyeongsangnam.gyeongnam_batch",
    "scrapers.gangwon.gangwon_batch",
    "scrapers.gongsa.gongsa_batch",
    "scrapers.jeju.jeju_do",
    "scrapers.jeju.jeju_si",
    "scrapers.jeju.jpdc",
]

def collect_working_urls() -> set[str]:
    working: set[str] = set()
    for mod_path in SCRAPER_MODULES:
        try:
            mod = importlib.import_module(mod_path)
            entries = mod.SCRAPERS if hasattr(mod, "SCRAPERS") else [(mod.SOURCE, mod.scrape)]
            for src, _ in entries:
                working.add(src.source_url)
        except Exception as e:
            print(f"  WARN: could not load {mod_path}: {e}", file=sys.stderr)
    return working


# ── known-fail patterns ────────────────────────────────────────────────────
# Each is a substring to check against the URL. If ANY matches → F.
FAIL_SUBSTRINGS: list[str] = [
    # JS-rendered portal/saeol boards (require JS execution)
    "/portal/saeol/gosi/list.do",
    "/saeol/gosi/list.do",
    # eGovFrame portals that JS-onclick every link
    "index.jeonbuk",
    "index.jeonnam",
    "index.gyeong",
    # known JS-loaded city pages
    "index.daegu.go.kr",
    "www.daegu.go.kr",
    # molit/env-ministry endpoints that block bare requests
    "molit.go.kr/srocm/",    # Busan-region molit — blocks bare requests
    "me.go.kr",              # env-ministry — blocked
    "mcee.go.kr",            # env-ministry (old domain) — blocked
    # open_content JS pages (no tables in initial HTML)
    "/open_content/ko/page.do",
    "/open_content/main/govern",
    # .web boards that return HTML with no <table>
    ".web/board/",
    # jsessionid-based session URLs (도로공사 etc.)
    "jsessionid",
    # LH 주택토지공사 — returns only 1 notice (not useful)
    "lh.or.kr",
    # kwater — XML-based
    "kwater.or.kr",
    # ex.co.kr — jsessionid session-based
    "ex.co.kr",
    # 서울시 JS-loaded boards
    "news_notice.do",
    # 부산지방국토관리청 — 404
    "/brocm/",
    # 진영 국토관리사무소 — 404
    "m_22196",
    # 인천도시공사 — HTTP 500
    "idc.or.kr",
    # 광주시청 JS-loaded
    "www.gwangju.go.kr",
    "gb.gwangju.go.kr",
    # 세종도시교통공사 — no tables, JS-loaded
    "seojongtransit.or.kr",
    # 경북: portal/saeol and various JS platforms
    "open.content/ko/page.do",
    # 경남: portal/saeol, .web, open_content no-tables
    "gyeongnam.go.kr/index.gyeong",
    # OfrAction.do — JS or unverified
    "OfrAction.do",
    # 전북: various JS patterns
    "익산지방국토관리청",
    "/irocm/",               #익산지방 molit
    # 전북: eminwon JS-onclick boards (some work, flag specific failing ones below)
    # 군산 고시공고 OfrAction
    "gunsan.go.kr/eminwon",
    # 전남: JS-rendered portals
    "jeonnam.go.kr/index.jeonnam",
    # 전남 specific JS/blocked
    "gwangyang.go.kr/www/",          # 광양시청 공지 — JS list
    "damyang.go.kr",                  # 담양군 — JS onclick
    "goheung.go.kr/portal/bbs",       # 고흥군 — JS onclick
    "hampyeong.go.kr",                # 함평군 — SSL error
    "yeongam.go.kr",                  # 영암군 — JS onclick
    "sinan.go.kr",                    # 신안군 — JS onclick
    # 충남: various failures
    "seocheon.go.kr/cop/bbs",         # 서천군 — 0 notices returned
    "hongseong.go.kr/portal/bbs",     # 홍성군 공지 — 0 notices
    "taean.go.kr/cop/bbs",            # 태안군 — JS onclick
    "asan.go.kr/eminwon",             # 아산 eminwon — JS
    "cheonan.go.kr/www",              # 천안시 공지 — JS loaded
    "gongju.go.kr",                   # 공주시 — title links are JS onclick
    "nonsan.go.kr/cop/bbs",           # 논산시 — unverified
    "gyeryong.go.kr",                 # 계룡시 — very small, skip
    # 충북: various
    "cb.go.kr/index",                 # 충북도청 — JS
    "cheongju.go.kr/index",           # 청주시 — JS portal
    "chungju.go.kr/index",            # 충주시 — JS
    "jecheon.go.kr/index",            # 제천시 — JS
    "jincheon.go.kr/portal",          # 진천군 — 0 notices
    "eumseong.go.kr/portal",          # 음성군 — 0 notices
    "goesan.go.kr",                   # 괴산군 — JS
    "danyang.go.kr",                  # 단양군 — JS
    # 경기: various
    "gg.go.kr/bbs/board",             # 경기도청 — JS
    "suwon.go.kr/www/cityMaster",     # 수원시 공지 — JS loaded
    "seongnam.go.kr",                 # 성남시 — JS
    "anyang.go.kr",                   # 안양시 — SSL error
    "bucheon.go.kr",                  # 부천시 — JS
    "ansan.go.kr",                    # 안산시 — JS
    "uijeongbu.go.kr",                # 의정부 — JS
    "siheung.go.kr",                  # 시흥시 — JS
    "gwangmyeong.go.kr",              # 광명시 — JS
    "gapyeong.go.kr",                 # 가평군 — JS
    "yeoncheon.go.kr",                # 연천군 — JS
    "pocheon.go.kr",                  # 포천시 — JS
    "dongducheon.go.kr",              # 동두천 — JS
    "yangpyeong.go.kr",               # 양평군 — JS
    "yeoju.go.kr",                    # 여주시 — JS
    "icheon.go.kr",                   # 이천시 — JS (some may work)
    "yongin.go.kr",                   # 용인시 — JS
    "gimpo.go.kr",                    # 김포시 — JS
    "gapyeong.go.kr",                 # 가평군 — JS
    "hanam.go.kr",                    # 하남시 — JS
    "gwacheon.go.kr",                 # 과천시 — JS
    "osan.go.kr",                     # 오산시 — JS
    "uiwang.go.kr",                   # 의왕시 — JS
    "yangju.go.kr",                   # 양주시 — JS
    # 강원: SSL / timeout / JS onclick
    "gn.go.kr",                       # 강원도청 — JS
    "gangneung.go.kr",                # 강릉시 — SSL error
    "sokcho.go.kr",                   # 속초시 — SSL error
    "donghae.go.kr",                  # 동해시 — SSL error (dh.go.kr alternate)
    "dh.go.kr",                       # 동해시 alt domain — SSL
    "samcheok.go.kr",                 # 삼척시 — JS onclick
    "wonju.go.kr",                    # 원주시 — JS
    "chuncheon.go.kr",                # 춘천시 — JS
    "hwacheon.go.kr",                 # 화천군 — JS
    "pyeongchang.go.kr",              # 평창군 — JS
    "jeongseon.go.kr",                # 정선군 — JS
    "yw.go.kr",                       # 양양군 — JS (note: yangyang.go.kr same)
    "yd21.go.kr",                     # 영동군 alt domain — SSL
    "hongcheon.go.kr",                # 홍천군 — JS onclick
    "inje.go.kr",                     # 인제군 — JS
    "goseong.go.kr/gw",               # 고성군(강원) — JS
    # 경북: most are JS-rendered portal/saeol or open_content
    "pohang.go.kr",                   # 포항시 — JS
    "gyeongju.go.kr",                 # 경주시 — JS
    "gimcheon.go.kr",                 # 김천시 — JS
    "gumi.go.kr",                     # 구미시 — JS
    "yeongju.go.kr",                  # 영주시 (domain) — JS
    "yc.go.kr/www",                   # 영천시 — JS
    "sangju.go.kr",                   # 상주시 — JS
    "mungyeong.go.kr",                # 문경시 — JS
    "gb.go.kr/Main/page.do?mnu_uid=3", # 경산시 — JS
    "gunwi.go.kr",                    # 군위군 — JS
    "uiseong.go.kr",                  # 의성군 — JS
    "cheongsong.go.kr",               # 청송군 — JS
    "cheongdo.go.kr",                 # 청도군 — JS
    "goryeong.go.kr",                 # 고령군 — JS
    "seongju.go.kr",                  # 성주군 — JS
    "chilgok.go.kr",                  # 칠곡군 — JS
    "bonghwa.go.kr",                  # 봉화군 — JS
    "uljin.go.kr",                    # 울진군 — JS
    "ulleung.go.kr",                  # 울릉군 — JS
    # 경남: portal/saeol JS, .web NO TABLES, etc.
    "changwon.go.kr",                 # 창원시 — JS
    "jinju.go.kr",                    # 진주시 — JS
    "tongyeong.go.kr",                # 통영시 — JS
    "sacheon.go.kr",                  # 사천시 — JS
    "gimhae.go.kr",                   # 김해시 — JS
    "miryang.go.kr",                  # 밀양시 — JS
    "geoje.go.kr",                    # 거제시 — JS
    "yangsan.go.kr",                  # 양산시 — JS
    "haman.go.kr",                    # 함안군 — JS
    "changnyeong.go.kr",              # 창녕군 — JS
    "namhae.go.kr",                   # 남해군 — JS
    "hadong.go.kr",                   # 하동군 — JS
    "sancheong.go.kr",                # 산청군 — JS
    "hamyang.go.kr",                  # 함양군 — JS
    "geochang.go.kr",                 # 거창군 — JS
    "hapcheon.go.kr",                 # 합천군 — JS
    "uiryeong.go.kr/portal",          # 의령군 portal/saeol — JS (plain board works)
    "goseong.go.kr/portal",           # 고성군(경남) portal/saeol — JS
    # 부산: failed/404
    "busan.go.kr/web/busanjunggu",    # 부산 중구 — JS
    "busanmpa.go.kr",                 # 부산항만공사 — blocked
    "bjinvest.or.kr",                 # 부산투자진흥공사 — JS
    # 대구: most JS-loaded
    "daeguenv.go.kr",                 # 대구환경청 — blocked
    "goryeong.go.kr",                 # already listed
    # 인천: HTTP 500
    "idco.or.kr",                     # 인천도시공사 — HTTP 500 (different from idc.or.kr)
    # 광주: JS-loaded
    "gwangju.go.kr/board",            # 광주시청 board — JS
    # national agencies known-fail
    "pblanNow.do",                    # 경기도 기술심사평가원 — JS
    "singisul.daegu.go.kr",           # 대구 신기술 — JS
    "calspia.go.kr",                  # 건설산업교육원 — JS
    "chungnam.go.kr/cnportal/province", # 충남 province page (not notices)
    # 제주
    "seogwipo.go.kr",                 # 서귀포시 — returns empty initial HTML
    "jeju.go.kr/news/news/law/jeju2", # 제주도청 고시 alt URL — connection reset
    # 세종
    "sctc.kr",                        # 세종도시교통공사 — no tables, JS-loaded
    # 경기 — portal/saeol boards
    "/portal/saeol/gosiList.do",
    "/portal/saeol/noticeList.do",
    # g2b.go.kr — national bidding portal (different system, not a simple HTML table board)
    "g2b.go.kr",
    # 고양시 입찰/보조금 — JS pages
    "goyang.go.kr/www/link/",
    # 강원도청 — portal JS-rendered
    "state.gwd.go.kr",                # 강원도청 포털 — JS
    # 강원 화천군 — contents.do (no table / JS)
    "ihc.go.kr",                      # 화천군 alt domain — no table
    # 충남/충북 — contents.do portal, BBSMSTR JS, saeolGosi JS
    "chungbuk.go.kr/www/contents.do", # 충북도청 — JS portal
    "cheongju.go.kr/www/contents.do", # 청주시 — JS portal
    "cheongju.go.kr/sangdang/",       # 청주 상당구 — JS portal
    "cheongju.go.kr/seowon/",         # 청주 서원구 — JS portal
    "cheongju.go.kr/heungdeok/",      # 청주 흥덕구 — JS portal
    "cheongju.go.kr/cheongwon/",      # 청주 청원구 — JS portal
    "jp.go.kr/kor/sub03",             # 증평군 — JS portal
    "hongseong.go.kr/prog/bbsArticle", # 홍성군 공지 BBSMSTR — JS
    "hongseong.go.kr/prog/saeolGosi", # 홍성군 고시 — saeolGosi JS
    "yesan.go.kr/bbs/BBSMSTR",        # 예산군 공지 BBSMSTR — JS
    "yesan.go.kr/prog/saeolGosi",     # 예산군 고시 — saeolGosi JS
    "brcn.go.kr/prog/eminwon",        # 보령시 고시 eminwon — JS
    "taean.go.kr/kor/sub02_03",       # 태안군 고시 — JS/no table
    "dangjin.go.kr/kor/sub03_02",     # 당진시 고시 — JS
    "nonsan.go.kr/kor/html",          # 논산시 — no table
    "seocheon.go.kr/kor/sub04",       # 서천군 일반공고 — no table
    "cheongyang.go.kr/kor/sub04",     # 청양군 고시 — no table
    "cheonan.go.kr/bbs/BBSMSTR",      # 천안시 공지 BBSMSTR — JS
    "ddc.go.kr/ddc/selectBbsNttList", # ddc.go.kr = 동두천시 (xlsx mislabeled)
    # 전북 — eminwon SSL, portal/saeol JS
    "eminwon.iksan.go.kr",            # 익산시 고시 eminwon — SSL
    "eminwon.jeongeup.go.kr",         # 정읍시 고시 eminwon — SSL
    "jinan.go.kr/board/list.jinan",   # 진안군 공지 — no table
    "jangsu.go.kr/board/list.jangsu", # 장수군 공지 — no table
    "gochang.go.kr/board/list.gochang", # 고창군 — no table
    "jbdc.co.kr",                     # 전북개발공사 — SSL
    # 전남 — various JS/no-table
    "suncheon.go.kr/kr/news/0004",    # 순천시 고시 — wrong page content
    "gokseong.go.kr/board/GosiList",  # 곡성군 고시 — JS
    "gurye.go.kr/board/GosiList",     # 구례군 고시 — JS
    "goheung.go.kr/boardList.do",     # 고흥군 공지 — no table
    "goheung.go.kr/contentsView.do",  # 고흥군 고시 — no table
    "shinan.go.kr",                   # 신안군 — wscms JS
    "haenam.go.kr/index.9is",         # 해남군 고시 — JS portal
    "cndc.kr",                        # 충남개발공사 — JS
    "jndc.co.kr",                     # 전남개발공사 — JS
    # 경북 — portal/bbs JS, page.do JS onclick
    "gc.go.kr/portal/bbs/list.do",    # 김천시 — portal/bbs JS
    "yc.go.kr/portal/bbs/list.do",    # 영천시 — portal/bbs JS
    "gbmg.go.kr/portal/bbs/list.do",  # 문경시 — portal/bbs JS
    "usc.go.kr/ko/page.do",           # 의성군 — page.do JS onclick
    "sj.go.kr/page.do",               # 성주군 — page.do JS onclick
    "cs.go.kr/news",                  # 청송군 — .web no table
    "gbdc.co.kr",                     # 경북개발공사 — JS
    # 경남 — .web no table
    "cng.go.kr",                      # 창녕군 — .web no table
    "hc.go.kr",                       # 합천군 — .web no table
    "gndc.co.kr",                     # 경남개발공사 — JS
    # 강원 — JS/timeout/SSL
    "yanggu.go.kr",                   # 양구군 — JS portal
    "yangyang.go.kr/gw/portal",       # 양양군 — JS portal
    # 인천
    "ih.co.kr",                       # 인천도시공사 — connection reset
    # 원주지방국토관리청 — connection reset
    "molit.go.kr/wrocm",
    # 창원시 통합데이터
    "changwon.go.kr/portal",
    # 경기 — JS onclick boards
    "hscity.go.kr/www/gosi/BD_notice",  # 화성시 — javascript:opGosiView onclick
    "gm.go.kr/pt/user/nftcBbs/",        # 광명시 — JS board
    "paju.go.kr/user/board/BD_board",   # 파주시 — JS onclick
    # 전북 — index.* JS portals
    "index.9is",
    "index.jinan",
    "index.jangsu",
    "index.imsil",
    "index.gochang",
]

# Additional: entire domains known to be SSL-error sites
FAIL_DOMAINS: set[str] = {
    "anyang.go.kr",
    "gn.go.kr",
    "gangneung.go.kr",
    "sokcho.go.kr",
    "dh.go.kr",
    "yd21.go.kr",
    "hampyeong.go.kr",
}


# Known path aliases: xlsx URL → scraper URL (when paths differ but content is the same)
PATH_ALIASES: dict[str, str] = {
    # 경남 도로관리사업소: xlsx uses /street/board/list.gyeong, scraper uses /board/list.gyeong
    "gyeongnam.go.kr/street/board/list.gyeong": "gyeongnam.go.kr/board/list.gyeong",
}


def normalize_url(url: str) -> str:
    """netloc/path — strips scheme, query, fragment; lowercases netloc"""
    try:
        p = urlparse(url)
        netloc = p.netloc.lstrip("www.").lower()
        norm = f"{netloc}{p.path}"
        return PATH_ALIASES.get(norm, norm)
    except Exception:
        return url


def classify(url: str, working: set[str], working_norm: set[str]) -> str:
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return ""

    # exact match first
    if url in working:
        return "T"

    # normalized match (handles http↔https, www prefix, minor path differences)
    norm = normalize_url(url)
    if norm in working_norm:
        return "T"

    # check domain against known SSL-fail set
    try:
        domain = urlparse(url).netloc.lstrip("www.")
        if domain in FAIL_DOMAINS or any(domain.endswith("." + d) for d in FAIL_DOMAINS):
            return "F"
    except Exception:
        pass

    # check fail substrings
    for pat in FAIL_SUBSTRINGS:
        if pat in url:
            return "F"

    return ""


def main():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path("links_v2.xlsx")
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found. Run from repo root.", file=sys.stderr)
        sys.exit(1)

    print("Collecting working URLs from scrapers…")
    working = collect_working_urls()
    print(f"  {len(working)} working source URLs found.")

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb["Sheet1"]

    # Write header in col F row 1
    ws.cell(row=1, column=6, value="상태")

    working_norm = {normalize_url(u) for u in working}

    counts = {"T": 0, "F": 0, "": 0}
    for row_idx in range(2, ws.max_row + 1):
        url_cell = ws.cell(row=row_idx, column=5)
        url = url_cell.value or ""
        status = classify(url, working, working_norm)
        ws.cell(row=row_idx, column=6, value=status)
        counts[status] += 1

    out_path = xlsx_path  # overwrite in-place
    wb.save(str(out_path))
    print(f"\nSaved → {out_path}")
    print(f"  T (working) : {counts['T']}")
    print(f"  F (failed)  : {counts['F']}")
    print(f"  blank       : {counts['']}")


if __name__ == "__main__":
    main()

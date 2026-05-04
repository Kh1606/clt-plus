"""
Parse list.xlsx → src/data/regions.json.

Two sheets:
  주요관청 (sheet1): metro-level — col B (도시) groups col C (기관) entries
  시-도   (sheet2): provinces  — col B (도시) groups col C (기관/시·군) entries
                    with merged-cell pattern (도시 only on first row of each group)

Each row's col D (페이지) is a page-type label; col E holds the real URL
as a hyperlink target. We carry the last-seen 도시/기관 forward across blank rows.
"""

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "list.xlsx"
OUT = ROOT / "src" / "data" / "regions.json"


def parse_sheet(ws, source_label: str):
    """Yield (region, sub_entity, page, url) tuples, carrying merged values forward."""
    cur_region = None
    cur_sub = None
    for i, row in enumerate(ws.iter_rows(min_row=3), start=3):
        # cells: A=순번 B=도시 C=기관 D=페이지 E=링크
        region_v = row[1].value
        sub_v = row[2].value
        page_v = row[3].value
        link_cell = row[4]

        if region_v:
            cur_region = str(region_v).strip()
        if sub_v:
            cur_sub = str(sub_v).strip()

        if not (cur_region and cur_sub and page_v and link_cell.hyperlink):
            continue

        url = link_cell.hyperlink.target
        if not url:
            continue

        yield (
            cur_region,
            cur_sub,
            str(page_v).strip(),
            url.strip(),
            source_label,
        )


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=False)

    # Group: region -> sub_entity -> [{page, url, source}]
    by_region = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for region, sub, page, url, src in parse_sheet(ws, sheet_name):
            sub_map = by_region.setdefault(region, {})
            sources = sub_map.setdefault(sub, [])
            sources.append({"page": page, "url": url, "sheet": src})

    # Convert to list-of-objects (UI-friendly)
    out = []
    for region, sub_map in by_region.items():
        sub_entities = [
            {"name": name, "sources": sources}
            for name, sources in sub_map.items()
        ]
        out.append({"region": region, "subEntities": sub_entities})

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    total_subs = sum(len(r["subEntities"]) for r in out)
    total_sources = sum(
        len(s["sources"]) for r in out for s in r["subEntities"]
    )
    print(f"Wrote {OUT.relative_to(ROOT)}")
    print(f"  regions:      {len(out)}")
    print(f"  sub-entities: {total_subs}")
    print(f"  sources:      {total_sources}")


if __name__ == "__main__":
    main()
